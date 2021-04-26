# Step 01: install openpyxl: pip install openpyxl

# Read the data from Excel

import openpyxl

# load the workbook
wb = openpyxl.load_workbook('data.xlsx')

# return all sheet names of given workbook
print(wb.sheetnames)

# return active sheet name of given workbook
print(wb.active)

# creating object from given sheet
sh = wb['data1']

# for rows count
print('rows count: ', sh.max_row)
# for columns count
print('columns count: ', sh.max_column)

# way 01
#for row in range(1, sh.max_row+1):
#    for column in range(1,  sh.max_column+1):
#        print(sh.cell(row, column).value)


# way 02: read the value using excel index
# print(sh['A1'].value)
for row in sh['A1' : 'D3']:
   for column in row:
       print(column.value)
   print()
